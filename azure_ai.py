from azure.ai.formrecognizer import DocumentAnalysisClient
from azure.core.credentials import AzureKeyCredential
import pandas as pd
from typing import List, Dict, Tuple
import os
from datetime import datetime
import re
from openpyxl.utils import get_column_letter

AZURE_KEY = os.getenv('AZURE_KEY', '')
AZURE_ENDPOINT = os.getenv('AZURE_ENDPOINT', '')

class KolmarCosmeticOCR:
    """콜마 화장품 제형 표 OCR 전용 클래스 (예외 사례 보완 완성)"""
    
    def __init__(self):
        """Azure Document Intelligence 클라이언트 초기화"""
        self.endpoint = AZURE_ENDPOINT
        self.key = AZURE_KEY
        
        self.client = DocumentAnalysisClient(
            endpoint=self.endpoint,
            credential=AzureKeyCredential(self.key)
        )
        
        print("✅ Azure Document Intelligence 연결 완료")
        print(f"📍 엔드포인트: {self.endpoint}")
    
    def _clean_checkbox_and_newline(self, value: str) -> str:
        """
        체크박스와 줄바꿈 제거 (개선)
        
        Phase, Code, 실험 ID, 모든 값에 적용
        """
        if not value:
            return ''
        
        value = str(value)
        
        # 체크박스 제거
        checkbox_words = [
            ':selected:', ':unselected:', ':checked:', ':unchecked:',
            ':SELECTED:', ':UNSELECTED:', ':CHECKED:', ':UNCHECKED:',  # 대문자 추가
            ':Selected:', ':Unselected:', ':Checked:', ':Unchecked:'   # 타이틀케이스 추가
        ]
        for checkbox_word in checkbox_words:
            value = value.replace(checkbox_word, '')
        
        # 줄바꿈 제거
        value = value.replace('\n', '').replace('\r', '').strip()
        
        return value
    
    def _normalize_experiment_value(self, value: str) -> str:
        """
        실험값 1차 정규화 (RULE 1~6)
        
        단계:
        1. 체크박스 제거
        2. 공백 정리
        3. 🆕 쉼표/콜론 → 점 변환
        4. X 정규화
        5. 소문자 → 대문자
        6. 잘못된 점 제거
        7. '=' 제거
        """
        if not value:
            return ''
        
        value = value.strip()
        
        # STEP 1: 체크박스 제거
        value = self._clean_checkbox_and_newline(value)
        value = value.strip()
        
        # STEP 2: TO100 특수 표현 유지
        if 'TO' in value.upper():
            return value
        
        # 🆕 STEP 3: 쉼표/콜론 → 점 변환 (숫자만)
        # 8,00 → 8.00
        # 5:00 → 5.00
        # 2,0 → 2.0
        if re.match(r'^\d+[,:]\d+$', value):
            value = value.replace(',', '.').replace(':', '.')
            print(f"    🔧 정규화: 쉼표/콜론 → 점 변환 → '{value}'")
        
        # STEP 4: X 변형 정규화
        x_variants = ['×', '✕', '✗', '*']
        if value in x_variants:
            value = 'X'
        
        # STEP 5: 소문자 x → 대문자 X
        if value.lower() == 'x':
            value = 'X'
        
        # STEP 6: 소수점이 여러 개면 마지막만 유지
        # 예: 1.2.3 → 1.23
        if value.count('.') > 1:
            parts = value.split('.')
            value = ''.join(parts[:-1]) + '.' + parts[-1]
        
        # STEP 7: 잘못된 점 제거
        # 10. → 10
        while value.endswith('.') and value.count('.') > 1:
            value = value[:-1]
        
        if value.endswith('.') and len(value) > 1 and value[:-1].replace('.', '').isdigit():
            value = value[:-1]
        
        # 🆕 STEP 8: '=' 제거
        value = value.replace('=', '').strip()
        
        return value
    
    def _validate_experiment_value(self, value: str) -> str:
        """
        실험값 2차 검증 (최종 보정)
        
        RULE 7: 숫자가 아니고 TO100도 아니면 텍스트 → '0'
        """
        if not value:
            return ''
        
        original = value  # ✅ 원본 저장
        value = value.strip()
        
        # TO100, TO 100 같은 특수 표현은 유지
        if 'TO' in value.upper():
            return value
        
        # 🆕 유럽식 소수점 (쉼표)
        # if re.match(r'^\d+,\d*$', value):  # 8,00 or 2,0
        #     return value
        
        # 🆕 시간 형식 (콜론)
        # if re.match(r'^\d+:\d+$', value):  # 2:0 or 5:00
        #     return value
        
        # 1) 순수 숫자: 10, 10.5, 0.5
        if re.match(r'^\d+\.?\d*$', value):
            return value
        
        # 2) 부등호 포함: <10, >5
        if re.match(r'^[<>≤≥]\s*\d+\.?\d*$', value):
            return value
        
        # 3) 범위: 5-10, 5~10
        if re.match(r'^\d+\.?\d*\s*[-~]\s*\d+\.?\d*$', value):
            return value
        
        # 4) 퍼센트: 10%, 5.5%
        if re.match(r'^\d+\.?\d*%$', value):
            return value
        
        # 5) 0 또는 0.0
        if value == '0' or value == '0.0':
            return value
        
        # 그 외 텍스트는 0으로 변환
        print(f"  ⚠️ RULE 7: 텍스트 발견 → '0' 변환: '{value}'")
        return '0'
    
    def _correct_phase(self, phase: str) -> str:
        """
        Phase 보정 (개선)
        
        RULE 6: Phase에는 알파벳만 존재
        - 체크박스 및 줄바꿈 제거
        - '1' → 'I'
        - '0' → 'O'
        """
        if not phase:
            return ''
        
        phase = phase.strip()
        
        # 체크박스 및 줄바꿈 제거 (개선)
        phase = self._clean_checkbox_and_newline(phase)
        
        # 숫자 → 알파벳 변환
        corrections = {
            '1': 'I',
            '0': 'O',
            'l': 'I',  # 소문자 L도 I로
            '8': 'B',  # 🆕 추가: 숫자 8 → 알파벳 B
        }
        
        for wrong, correct in corrections.items():
            phase = phase.replace(wrong, correct)
        
        return phase.upper()
    
    def _detect_empty_columns(self, ingredients: List[Dict], experiment_cols: List[str]) -> List[str]:
        """
        빈 시험 컬럼 감지
        
        RULE 8: 모든 원료에서 값이 없는 컬럼 찾기
        
        Args:
            ingredients: 원료 리스트
            experiment_cols: 실험 컬럼 리스트
        
        Returns:
            빈 컬럼 리스트
        """
        empty_cols = []
        
        for exp_col in experiment_cols:
            # 해당 컬럼의 모든 값이 빈칸인지 확인
            all_empty = True
            for ingredient in ingredients:
                value = ingredient.get(exp_col, '').strip()
                if value:  # 값이 하나라도 있으면
                    all_empty = False
                    break
            
            if all_empty:
                empty_cols.append(exp_col)
        
        if empty_cols:
            print(f"\n🔍 RULE 8: 빈 시험 컬럼 감지: {empty_cols}")
        
        return empty_cols
    
    def _apply_data_correction_rules(self, ingredients: List[Dict], experiment_cols: List[str]) -> List[Dict]:
        """
        데이터 보정 룰 적용
        
        RULE 1: 첫번째 실험 컬럼에서 공란 → '0'
        RULE 2: '-' → '0', X/x → '0', 체크박스 제거
        RULE 3: 두번째 이후 컬럼에서 공란 → 이전 컬럼 값 복사 (빈 컬럼 건너뛰기)
        RULE 4: Phase 공란 → 이전 Phase 상속
        RULE 5: 원료 코드 없는 행 삭제
        RULE 6: Phase 보정 (1→I, 0→O)
        RULE 7: 텍스트 → '0' (TO100 제외)
        RULE 8: 빈 시험 컬럼 감지 및 건너뛰기
        """
        
        print("\n🔧 데이터 보정 룰 적용 중...")
        
        if not experiment_cols:
            return ingredients
        
        # RULE 8: 빈 컬럼 감지
        empty_cols = self._detect_empty_columns(ingredients, experiment_cols)
        
        # RULE 4: Phase 공란 → 이전 Phase 상속
        prev_phase = ''
        
        for ingredient in ingredients:
            
            # ============================================
            # 🆕 추가 1: 보정 플래그 딕셔너리 초기화
            # ============================================
            correction_flags = {}
            # RULE 6: Phase 보정
            if 'Phase' in ingredient:
                original_phase = ingredient['Phase']
                corrected_phase = self._correct_phase(original_phase)
                if original_phase != corrected_phase:
                    ingredient['Phase'] = corrected_phase
                    print(f"  RULE 6: Phase 보정 '{original_phase}' → '{corrected_phase}'")
            
            # RULE 4: Phase 공란 시 상속
            if not ingredient.get('Phase', '').strip():
                ingredient['Phase'] = prev_phase
            else:
                prev_phase = ingredient['Phase']
            
            code = None
            for key in ingredient.keys():
                if key.lower() == 'code':
                    code = ingredient[key]
                    break
            
            if not code:
                continue
            
            # RULE 1, 3 (고도화): 실험값 보정
            for idx, exp_col in enumerate(experiment_cols):
                current_value = ingredient.get(exp_col, '').strip()
                
                # RULE 1: 첫 번째 컬럼 공란 → '0'
                if idx == 0:
                    if not current_value:
                        ingredient[exp_col] = '0'
                        correction_flags[exp_col] = 'filled_zero'
                        print(f"  RULE 1: [{code}] {exp_col} 공란 → '0'")
                
                # RULE 3 (고도화): 두 번째 이후 컬럼 공란 → 유효한 이전 값 복사
                else:
                    if not current_value:
                        # 현재 컬럼이 빈 컬럼이면 건너뛰기
                        if exp_col in empty_cols:
                            continue
                        
                        # 유효한 이전 컬럼 찾기 (빈 컬럼 제외하고 역방향 검색)
                        prev_value = None
                        source_col = None
                        
                        for prev_idx in range(idx - 1, -1, -1):
                            prev_col = experiment_cols[prev_idx]
                            
                            # 빈 컬럼이면 건너뛰기
                            if prev_col in empty_cols:
                                continue
                            
                            # 값이 있으면 사용
                            prev_value = ingredient.get(prev_col, '').strip()
                            if prev_value:
                                source_col = prev_col
                                ingredient[exp_col] = prev_value
                                correction_flags[exp_col] = 'copied'
                                print(f"  RULE 3: [{code}] {exp_col} 공란 → '{prev_value}' (from {source_col})")
                                break
            
            # RULE 7: 최종 텍스트 검증
            for exp_col in experiment_cols:
                # 빈 컬럼은 검증 제외
                if exp_col in empty_cols:
                    continue
                
                current_value = ingredient.get(exp_col, '').strip()
                if current_value:
                    # 🆕 주석: 이 시점에서는 이미 정규화된 값 (쉼표→점 변환 완료)
                    validated_value = self._validate_experiment_value(current_value)
                    if validated_value != current_value:
                        ingredient[exp_col] = validated_value
                        
            ingredient['_corrections'] = correction_flags
        print("✅ 데이터 보정 룰 적용 완료")
        
        return ingredients
    
    def _extract_from_meta_table(self, table, field_type: str) -> str:
        """
        메타데이터 테이블에서 정보 추출
        
        개선사항:
        - Formula No 라벨의 바로 다음 셀만 확인
        - ORIGINS 등 다른 라벨의 값 제외
        - 제품명에서 'No /', 'Date /' 제거
        """
        
        print(f"\n🔍 메타 테이블 추출 시도: {field_type}")
        print(f"  테이블 크기: {table.row_count}행 x {table.column_count}열")
        
        # 디버깅: 모든 셀 내용 출력
        print(f"  테이블 내용:")
        cells_by_row = {}
        for cell in table.cells:
            row_idx = cell.row_index
            if row_idx not in cells_by_row:
                cells_by_row[row_idx] = []
            cells_by_row[row_idx].append((cell.column_index, cell.content.strip()))
        
        for row_idx in sorted(cells_by_row.keys()):
            row_content = ' | '.join([f"[{col}]{content[:30]}" for col, content in sorted(cells_by_row[row_idx])])
            print(f"    행 {row_idx}: {row_content}")
        
        # 추출 로직
        for cell in table.cells:
            content = cell.content.strip()
            content_upper = content.upper().replace(' ', '')
            
            if field_type == 'formula_number':
                # 🔧 수정: Formula No/Formelle No 라벨을 정확히 찾기
                if ('FORMULANO' in content_upper or 
                    'FORMELLENO' in content_upper or  # OCR 오류
                    '처방번호' in content):
                    
                    print(f"    라벨 발견: '{content}' (행{cell.row_index}, 열{cell.column_index})")
                    
                    # 🔧 핵심: 바로 다음 셀(column_index + 1)만 확인
                    for next_cell in table.cells:
                        if (next_cell.row_index == cell.row_index and 
                            next_cell.column_index == cell.column_index + 1):
                            
                            value = next_cell.content.strip()
                            match = re.search(r'WE\d{4}', value.upper())
                            if match:
                                result = match.group()
                                print(f"  ✅ 문서번호 발견: '{result}' (셀: 행{cell.row_index}, 열{next_cell.column_index})")
                                return result
            
            elif field_type == 'product_name':
                # 제품 명 찾기
                if '제품' in content and '명' in content:
                    print(f"    라벨 발견: '{content}' (행{cell.row_index}, 열{cell.column_index})")
                    
                    # 같은 행의 다음 셀들 병합
                    values = []
                    for next_cell in sorted([c for c in table.cells 
                                        if c.row_index == cell.row_index and c.column_index > cell.column_index], 
                                        key=lambda x: x.column_index):
                        next_value = next_cell.content.strip()
                        
                        # 🔧 수정: 불필요한 텍스트 필터링 강화
                        if next_value and next_value not in ['DATE', 'Date', 'NO', 'No', '/', '']:
                            # Date, No 단어 제거
                            next_value = re.sub(r'\s*Date\s*/?\s*', '', next_value, flags=re.IGNORECASE)
                            next_value = re.sub(r'\s*No\s*/?\s*$', '', next_value, flags=re.IGNORECASE)
                            next_value = next_value.strip()
                            
                            if next_value:
                                values.append(next_value)
                    
                    if values:
                        result = ' '.join(values)
                        print(f"  ✅ 제품명 발견: '{result}' (행{cell.row_index})")
                        return result
            
            elif field_type == 'characteristics':
                if '처방특성' in content or '특성' in content:
                    print(f"    라벨 발견: '{content}' (행{cell.row_index}, 열{cell.column_index})")
                    
                    # 같은 행의 다음 셀들 병합
                    values = []
                    for next_cell in sorted([c for c in table.cells 
                                        if c.row_index == cell.row_index and c.column_index > cell.column_index], 
                                        key=lambda x: x.column_index):
                        next_value = next_cell.content.strip()
                        if next_value:
                            values.append(next_value)
                    
                    if values:
                        result = ' '.join(values)
                        print(f"  ✅ 처방특성 발견: '{result}' (행{cell.row_index})")
                        return result
        
        print(f"  ⚠️ {field_type} 추출 실패")
        return ''

    def extract_cosmetic_formula_table(self, image_path: str) -> Dict:
        """화장품 제형 실험 표 추출"""
        print(f"\n🔍 이미지 분석 시작: {os.path.basename(image_path)}")
        
        with open(image_path, 'rb') as f:
            image_data = f.read()
        
        print("📊 테이블 구조 분석 중...")
        poller = self.client.begin_analyze_document("prebuilt-layout", document=image_data)
        result = poller.result()
        
        print(f"📋 감지된 테이블 수: {len(result.tables)}")
        for idx, tbl in enumerate(result.tables):
            print(f"  테이블 {idx}: {tbl.row_count}행 x {tbl.column_count}열")
        
        # ========== 메타데이터 추출: 3단계 전략 ==========
        document_info = {
            'formula_number': '',
            'product_name': '',
            'characteristics': ''
        }
        
        if len(result.tables) >= 2:
            table_sizes = [(idx, tbl.row_count * tbl.column_count) for idx, tbl in enumerate(result.tables)]
            table_sizes.sort(key=lambda x: x[1])
            
            small_idx = table_sizes[0][0]
            large_idx = table_sizes[-1][0]
            
            print(f"  → 작은 테이블(메타): 테이블 {small_idx}")
            print(f"  → 큰 테이블(제형): 테이블 {large_idx}")
            
            # 1단계: 작은 테이블(메타)에서 추출
            meta_table = result.tables[small_idx]
            formula_number = self._extract_from_meta_table(meta_table, 'formula_number')
            product_name = self._extract_from_meta_table(meta_table, 'product_name')
            characteristics = self._extract_from_meta_table(meta_table, 'characteristics')
            
            document_info['formula_number'] = formula_number
            document_info['product_name'] = product_name
            document_info['characteristics'] = characteristics
            
            # 2단계: 제형 테이블 상단에서 추출 (부족한 정보 보완)
            if not document_info['product_name'] or not document_info['characteristics']:
                print(f"\n⚠️ 메타 테이블에서 일부 정보 추출 실패, 제형 테이블 상단 확인")
                table = result.tables[large_idx]
                formula_header_info = self._extract_from_formula_table_header(table)
                
                if not document_info['formula_number']:
                    document_info['formula_number'] = formula_header_info['formula_number']
                if not document_info['product_name']:
                    document_info['product_name'] = formula_header_info['product_name']
                    print(f"  🔄 제품명 (제형 테이블): '{document_info['product_name']}'")
                if not document_info['characteristics']:
                    document_info['characteristics'] = formula_header_info['characteristics']
                    print(f"  🔄 처방특성 (제형 테이블): '{document_info['characteristics']}'")
            
            # 3단계: 전체 텍스트 fallback
            if not document_info['formula_number'] or not document_info['product_name']:
                print(f"\n⚠️ 여전히 정보 부족, 전체 텍스트에서 재시도")
                full_text = result.content
                fallback_info = self._extract_document_info(full_text)
                
                if not document_info['formula_number']:
                    document_info['formula_number'] = fallback_info.get('formula_number', 'Unknown')
                    print(f"  🔄 문서번호 (전체 텍스트): '{document_info['formula_number']}'")
                
                if not document_info['product_name']:
                    document_info['product_name'] = fallback_info.get('product_name', '제품명 미확인')
                    print(f"  🔄 제품명 (전체 텍스트): '{document_info['product_name']}'")
            
            table = result.tables[large_idx]
            
        else:
            # 테이블 1개면 전체 텍스트에서 추출
            full_text = result.content
            document_info = self._extract_document_info(full_text)
            table = result.tables[0]
        
        print(f"📋 문서번호: {document_info.get('formula_number', 'Unknown')}")
        print(f"📦 제품명: {document_info.get('product_name', 'Unknown')}")
        
        if not result.tables:
            print("❌ 테이블을 찾을 수 없습니다.")
            return {}
        
        print(f"✅ 제형 테이블 선택: {table.row_count}행 x {table.column_count}열")
        
        raw_table_data = self._extract_raw_table(table)
        formula_data = self._parse_and_clean_table(table)
        formula_data.update(document_info)
        formula_data['raw_table'] = raw_table_data
        
        print(f"🧴 추출된 원료 수: {len(formula_data.get('ingredients', []))}개")
        
        return formula_data
    
    def _extract_from_formula_table_header(self, table) -> Dict:
        """
        제형 테이블 상단에서 메타데이터 추출
        
        많은 경우 제형 테이블의 처음 2-3행에 제품명, 처방특성 등이 있음
        """
        print(f"\n🔍 제형 테이블 상단에서 메타데이터 추출 시도")
        
        info = {
            'formula_number': '',
            'product_name': '',
            'characteristics': ''
        }
        
        # 처음 5행만 확인
        cells_by_row = {}
        for cell in table.cells:
            if cell.row_index < 5:  # 처음 5행만
                row_idx = cell.row_index
                if row_idx not in cells_by_row:
                    cells_by_row[row_idx] = {}
                cells_by_row[row_idx][cell.column_index] = cell.content.strip()
        
        # 각 행 확인
        for row_idx in sorted(cells_by_row.keys()):
            row_data = cells_by_row[row_idx]
            
            for col_idx, content in row_data.items():
                content_lower = content.lower()
                
                # 제품명 찾기
                if '제품' in content and '명' in content:
                    # 같은 행의 다음 셀들 병합
                    values = []
                    for next_col in sorted([c for c in row_data.keys() if c > col_idx]):
                        next_value = row_data[next_col]
                        if next_value and next_value not in ['DATE', 'Date', 'NO', '/', '']:
                            if 'DATE' in next_value or 'Date' in next_value:
                                next_value = next_value.split('DATE')[0].split('Date')[0].strip()
                            if next_value:
                                values.append(next_value)
                    
                    if values:
                        info['product_name'] = ' '.join(values)
                        print(f"  ✅ 제품명 발견: '{info['product_name']}' (행{row_idx})")
                
                # 처방특성 찾기
                if '처방특성' in content or ('처방' in content and '특성' in content):
                    # 같은 행의 다음 셀들 병합
                    values = []
                    for next_col in sorted([c for c in row_data.keys() if c > col_idx]):
                        next_value = row_data[next_col]
                        if next_value:
                            values.append(next_value)
                    
                    if values:
                        info['characteristics'] = ' '.join(values)
                        print(f"  ✅ 처방특성 발견: '{info['characteristics']}' (행{row_idx})")
                
                # Formula No 찾기
                if 'formula' in content_lower or 'WE' in content.upper():
                    match = re.search(r'WE\d{4}', content.upper())
                    if match:
                        info['formula_number'] = match.group()
                        print(f"  ✅ 문서번호 발견: '{info['formula_number']}' (행{row_idx})")
        
        return info

    def _extract_raw_table(self, table) -> pd.DataFrame:
        """원본 테이블 추출"""
        table_matrix = {}
        for cell in table.cells:
            row_idx = cell.row_index
            col_idx = cell.column_index
            
            if row_idx not in table_matrix:
                table_matrix[row_idx] = {}
            
            table_matrix[row_idx][col_idx] = cell.content.strip()
        
        # 모든 행의 최대 컬럼 수 찾기
        max_cols = 0
        for row_data in table_matrix.values():
            if row_data:
                max_cols = max(max_cols, max(row_data.keys()) + 1)
        
        # 모든 행을 동일한 컬럼 수로 맞추기
        rows_data = []
        for row_idx in sorted(table_matrix.keys()):
            row = []
            for col_idx in range(max_cols):
                row.append(table_matrix[row_idx].get(col_idx, ''))
            rows_data.append(row)
        
        # 컬럼명 생성
        columns = [f'Col_{i}' for i in range(max_cols)]
        
        # DataFrame 생성
        df = pd.DataFrame(rows_data, columns=columns)
        
        return df
    
    def _extract_document_info(self, text: str) -> Dict:
        """문서 정보 추출 (개선)"""
        info = {}
        
        # 🔧 문서번호: WE + 4자리 숫자
        formula_match = re.search(r'WE\d{4}', text.upper())
        info['formula_number'] = formula_match.group() if formula_match else 'Unknown'
        
        # Origin (부차적 정보)
        origin_match = re.search(r'Origin[:\s]*([A-Z0-9]+)', text, re.IGNORECASE)
        info['origin'] = origin_match.group(1) if origin_match else ''
        
        # 🔧 제품명: 여러 패턴 시도
        product_patterns = [
            r'제품\s*명[:\s]*([가-힣\s\w\(\)]+?)(?:DATE|Date|ORIGIN|Origin|\n|$)',  # 제품 명: XXX
            r'(?:페이스|에센스|세럼|크림|로션|토너)[가-힣\s\w\(\)]+제형',  # XXX 제형
            r'[가-힣]{2,}\s+[가-힣]{2,}\s+제형',  # 두 단어 이상 + 제형
        ]
        
        for pattern in product_patterns:
            product_match = re.search(pattern, text)
            if product_match:
                product_name = product_match.group(1) if product_match.lastindex else product_match.group()
                product_name = product_name.strip()
                
                # 불필요한 단어 제거
                for remove_word in ['DATE', 'Date', 'ORIGIN', 'Origin', '제품명', '제품 명']:
                    product_name = product_name.replace(remove_word, '')
                
                product_name = product_name.strip()
                
                if len(product_name) > 3:  # 최소 길이 체크
                    info['product_name'] = product_name
                    break
        
        if 'product_name' not in info:
            info['product_name'] = '제품명 미확인'
        
        # 처방특성
        characteristics_match = re.search(r'처방특성[:\s]*([가-힣\s\w\(\)]+)', text)
        info['characteristics'] = characteristics_match.group(1).strip() if characteristics_match else ''
        
        return info
    
    def _find_header_rows(self, table_matrix: Dict) -> Tuple[int, int]:
        """
        헤더 행 찾기 (개선: RAW MATERIALS가 이전/다음 행에 있는 경우 모두 처리)
        """
        main_header_row = None
        exp_id_row = None
        
        print(f"\n🔍 헤더 검색 중 (총 {len(table_matrix)}행)...")
        
        for row_idx in range(min(15, len(table_matrix))):
            if row_idx not in table_matrix:
                continue
            
            row_data = table_matrix[row_idx]
            row_text = ' '.join(str(v) for v in row_data.values()).upper()
            
            print(f"  행 {row_idx}: {row_text[:100]}...")
            
            if main_header_row is None:
                has_phase = any(keyword in row_text for keyword in ['PHASE', '상', 'STAGE'])
                has_code = any(keyword in row_text for keyword in ['CODE', '코드', '원료코드'])
                has_material = any(keyword in row_text for keyword in ['MATERIAL', '원료', 'RAW', '원료명'])
                
                # 🔥 수정: CODE만 있어도 이전/다음 행 확인
                if has_code:
                    # 🆕 이전 행에 MATERIAL 확인
                    prev_row_idx = row_idx - 1
                    has_material_prev = False
                    if prev_row_idx >= 0 and prev_row_idx in table_matrix:
                        prev_row_text = ' '.join(str(v) for v in table_matrix[prev_row_idx].values()).upper()
                        has_material_prev = any(keyword in prev_row_text for keyword in ['MATERIAL', '원료', 'RAW', '원료명'])
                    
                    # 다음 행에 MATERIAL 확인
                    next_row_idx = row_idx + 1
                    has_material_next = False
                    if next_row_idx in table_matrix:
                        next_row_text = ' '.join(str(v) for v in table_matrix[next_row_idx].values()).upper()
                        has_material_next = any(keyword in next_row_text for keyword in ['MATERIAL', '원료', 'RAW', '원료명'])
                    
                    # 🔧 수정: 현재/이전/다음 행 중 하나라도 MATERIAL 있으면 OK
                    if has_material or has_material_prev or has_material_next:
                        main_header_row = row_idx
                        print(f"✅ 메인 헤더 행: {row_idx} (CODE 발견)")
                        
                        if has_material_prev:
                            print(f"  ℹ️ RAW MATERIALS는 이전 행 {prev_row_idx}에 위치")
                        elif has_material_next:
                            print(f"  ℹ️ RAW MATERIALS는 다음 행 {next_row_idx}에 위치")
                        
                        # 🔥 수정: 실험 ID 행 찾기 (MATERIAL 위치에 따라 분기)
                        if has_material_prev:
                            # RAW MATERIALS가 이전 행이면, CODE 다음 행이 실험 ID
                            exp_id_row = row_idx + 1
                        elif has_material_next:
                            # RAW MATERIALS가 다음 행이면, 그 다음 행이 실험 ID
                            exp_id_row = row_idx + 2
                        else:
                            # RAW MATERIALS가 같은 행이면, 다음 행이 실험 ID
                            exp_id_row = row_idx + 1
                        
                        # 실험 ID 행 검증
                        if exp_id_row in table_matrix:
                            exp_row_data = table_matrix[exp_id_row]
                            
                            single_letters = []
                            for col_idx, value in exp_row_data.items():
                                cleaned = self._clean_checkbox_and_newline(str(value))
                                # 🆕 특수문자 제거 (H- → H)
                                cleaned = cleaned.replace('-', '').replace('_', '').strip()
                                
                                if cleaned and len(cleaned) == 1 and cleaned.isalpha():
                                    single_letters.append(cleaned)
                            
                            print(f"  실험 ID 행({exp_id_row}) 단일 알파벳: {single_letters}")
                            
                            if len(single_letters) >= 3:
                                print(f"✅ 실험 ID 행: {exp_id_row}")
                            else:
                                # 단일 알파벳이 부족하면 다음 행 시도
                                exp_id_row_alt = exp_id_row + 1
                                if exp_id_row_alt in table_matrix:
                                    exp_row_data_alt = table_matrix[exp_id_row_alt]
                                    single_letters_alt = []
                                    
                                    for col_idx, value in exp_row_data_alt.items():
                                        cleaned = self._clean_checkbox_and_newline(str(value))
                                        cleaned = cleaned.replace('-', '').replace('_', '').strip()
                                        if cleaned and len(cleaned) == 1 and cleaned.isalpha():
                                            single_letters_alt.append(cleaned)
                                    
                                    if len(single_letters_alt) >= 3:
                                        exp_id_row = exp_id_row_alt
                                        print(f"  ℹ️ 실험 ID를 다음 행 {exp_id_row}에서 발견: {single_letters_alt}")
                                        print(f"✅ 실험 ID 행: {exp_id_row}")
                        
                        break
                
                # 🎯 기존 로직: PHASE + CODE + MATERIAL이 모두 있으면 (호환성 유지)
                elif has_phase and has_code and has_material:
                    main_header_row = row_idx
                    print(f"✅ 메인 헤더 행: {row_idx} (PHASE + CODE + MATERIAL 발견)")
                    
                    # 다음 행이 실험 ID 행인지 확인
                    next_row_idx = row_idx + 1
                    if next_row_idx in table_matrix:
                        next_row_data = table_matrix[next_row_idx]
                        
                        single_letters = []
                        for col_idx, value in next_row_data.items():
                            cleaned = str(value).strip()
                            for checkbox in [':selected:', ':unselected:', ':checked:', ':unchecked:']:
                                cleaned = cleaned.replace(checkbox, '')
                            cleaned = cleaned.replace('\n', '').replace('\r', '').strip()
                            
                            if cleaned and len(cleaned) == 1 and cleaned.isalpha():
                                single_letters.append(cleaned)
                        
                        print(f"  다음 행 {next_row_idx}의 단일 알파벳: {single_letters}")
                        
                        if len(single_letters) >= 3:
                            exp_id_row = next_row_idx
                            print(f"✅ 실험 ID 행: {next_row_idx}")
                    break
        
        if main_header_row is None:
            print("\n⚠️ 헤더를 찾지 못했습니다.")
            print("💡 첫 5행 샘플:")
            for row_idx in range(min(5, len(table_matrix))):
                if row_idx in table_matrix:
                    sample_text = ' | '.join(str(v) for v in list(table_matrix[row_idx].values())[:5])
                    print(f"   행 {row_idx}: {sample_text[:100]}")
            
            print("\n⚠️ 첫 번째 행을 헤더로 사용합니다.")
            main_header_row = 0
            exp_id_row = 1 if 1 in table_matrix else None
        
        # ✅ 추가: 실험 ID 행 전체 출력 (디버깅용)
        if exp_id_row is not None and exp_id_row in table_matrix:
            print(f"\n📋 실험 ID 행({exp_id_row}) 전체 데이터:")
            exp_row_data = table_matrix[exp_id_row]
            for col_idx in sorted(exp_row_data.keys()):
                value = exp_row_data[col_idx]
                cleaned = self._clean_checkbox_and_newline(value)
                print(f"  Col_{col_idx}: '{value}' → '{cleaned}'")
        
        return main_header_row, exp_id_row
    
    def _align_raw_materials_header(self, table_matrix: Dict, header_row: int) -> Dict:
        """
        RAW MATERIALS 헤더를 실제 데이터 위치로 정렬 (전처리)
        
        문제: 헤더(Col_3)와 실제 데이터(Col_2) 위치 불일치
        해결: 헤더를 데이터가 있는 Col_2로 이동
        """
        if header_row not in table_matrix:
            return table_matrix
        
        header_data = table_matrix[header_row]
        
        # CODE 컬럼 찾기
        code_col = None
        for col_idx, value in header_data.items():
            if 'CODE' in str(value).upper():
                code_col = col_idx
                break
        
        # RAW MATERIALS 컬럼 찾기
        raw_mat_col = None
        for col_idx, value in header_data.items():
            if 'RAW' in str(value).upper() or 'MATERIAL' in str(value).upper():
                raw_mat_col = col_idx
                break
        
        if code_col is None or raw_mat_col is None:
            return table_matrix
        
        # 실제 데이터가 있는 컬럼 찾기 (CODE 다음 컬럼부터 확인)
        data_col = None
        max_data_count = 0
        
        for check_col in range(code_col + 1, raw_mat_col + 1):
            data_count = 0
            for check_row in range(header_row + 2, min(header_row + 20, len(table_matrix))):
                if check_row in table_matrix and check_col in table_matrix[check_row]:
                    cell_value = str(table_matrix[check_row][check_col]).strip()
                    if cell_value and cell_value not in ['nan', 'None', '']:
                        data_count += 1
            
            if data_count > max_data_count:
                max_data_count = data_count
                data_col = check_col
        
        # 헤더 정렬
        if data_col is not None and data_col != raw_mat_col:
            print(f"\n🔧 전처리: RAW MATERIALS 헤더 정렬")
            print(f"  Col_{raw_mat_col} → Col_{data_col}")
            table_matrix[header_row][data_col] = 'RAW MATERIALS'
            if raw_mat_col != data_col:
                table_matrix[header_row][raw_mat_col] = ''
            print(f"  ✅ 완료")
        
        return table_matrix
    
    def _identify_columns(self, table_matrix: Dict, header_row: int, exp_id_row: int = None) -> Dict:
        """컬럼 식별 (실험 컬럼 조건 강화 버전)"""
        if header_row not in table_matrix:
            print(f"⚠️ 헤더 행 {header_row}이 존재하지 않습니다.")
            return {}
        
        row_data = table_matrix[header_row]
        phase_col = None
        code_col = None
        name_col = None
        
        # 🔧 수정: exp_id_row를 파라미터로 받음
        if exp_id_row is None:
            exp_id_row = header_row + 1
        
        print(f"\n🔍 컬럼 식별 중 (헤더 행 {header_row}, 실험 ID 행 {exp_id_row}):")
        
        # 현재 행에서 컬럼 찾기
        for col_idx, value in row_data.items():
            value_upper = str(value).upper().strip()
            print(f"  Col_{col_idx}: '{value}' (upper: '{value_upper}')")
            
            if phase_col is None:
                if any(k in value_upper for k in ['PHASE', '상', 'STAGE']):
                    phase_col = col_idx
                    print(f"    ✅ Phase 컬럼 발견")
            
            if code_col is None:
                if any(k in value_upper for k in ['CODE', '코드', '원료코드']):
                    code_col = col_idx
                    print(f"    ✅ Code 컬럼 발견")
            
            if name_col is None:
                if any(k in value_upper for k in ['MATERIAL', '원료', 'RAW', '원료명', 'NAME']):
                    name_col = col_idx
                    print(f"    ✅ Name 컬럼 발견")
        
        # Phase가 없으면 이전 행에서 찾기
        if phase_col is None:
            prev_row_idx = header_row - 1
            if prev_row_idx >= 0 and prev_row_idx in table_matrix:
                prev_row_data = table_matrix[prev_row_idx]
                print(f"\n  ℹ️ Phase를 이전 행 {prev_row_idx}에서 검색:")
                
                for col_idx, value in prev_row_data.items():
                    value_upper = str(value).upper().strip()
                    if any(k in value_upper for k in ['PHASE', '상', 'STAGE']):
                        phase_col = col_idx
                        print(f"    ✅ Phase 컬럼 발견: Col_{col_idx} (이전 행)")
                        break
        
        print(f"\n📋 기본 컬럼 - Phase: {phase_col}, Code: {code_col}, Name: {name_col}")
        
        # 기본 컬럼이 없으면 기본값 설정
        if phase_col is None or code_col is None or name_col is None:
            print(f"⚠️ 기본 컬럼을 찾지 못했습니다!")
            print(f"💡 대안: 컬럼 인덱스 수동 설정 (Phase=0, Code=1, Name=2)")
            
            if phase_col is None:
                phase_col = 0
                print(f"   Phase를 Col_0으로 가정")
            if code_col is None:
                code_col = 1
                print(f"   Code를 Col_1로 가정")
            if name_col is None:
                name_col = 2
                print(f"   Name를 Col_2로 가정")
        
        # 🎯 실험 컬럼 찾기
        max_col = 0
        for row_idx, row in table_matrix.items():
            if row:
                row_max = max(row.keys())
                if row_max > max_col:
                    max_col = row_max
        
        # 🔥🔥🔥 핵심 수정: 이 줄을 삭제! 🔥🔥🔥
        # exp_id_row = header_row + 1  # ❌ 삭제
        # exp_id_row는 이미 파라미터로 받았으므로 재할당 금지!
                    
        experiment_cols = []
        
        print(f"\n🔬 실험 컬럼 찾기 시작:")
        print(f"  max_col = {max_col}")
        print(f"  실험 ID 행: {exp_id_row}")
        print(f"  제외: [Phase={phase_col}, Code={code_col}, Name={name_col}]")
        print(f"  확인 범위: Col_0 ~ Col_{max_col}")
        print(f"  행 범위: {exp_id_row} ~ {min(header_row + 20, len(table_matrix)) - 1}")
        
        for col_idx in range(max_col + 1):
            # Phase, Code, Name 컬럼은 제외
            if col_idx in [phase_col, code_col, name_col]:
                continue
                
            print(f"\n  Col_{col_idx} 확인 중...")
            
            # ✅ 추가: 실제 데이터 샘플 출력 (처음 5개)
            print(f"    === 실제 데이터 샘플 ===")
            sample_count = 0
            for check_row_idx in range(exp_id_row, min(header_row + 20, len(table_matrix))):
                if check_row_idx in table_matrix and col_idx in table_matrix[check_row_idx]:
                    cell_value = str(table_matrix[check_row_idx][col_idx]).strip()
                    if cell_value and cell_value not in ['nan', 'None', '']:
                        print(f"      행 {check_row_idx}: '{cell_value[:30]}'")
                        sample_count += 1
                        if sample_count >= 5:
                            break
                
            # ========== 🔥 1단계: 실험 ID 행에 단일 알파벳 확인 ==========
            has_experiment_id = False
            experiment_id_value = None

            if exp_id_row in table_matrix and col_idx in table_matrix[exp_id_row]:
                id_value = self._clean_checkbox_and_newline(str(table_matrix[exp_id_row][col_idx]))
                print(f"    실험 ID 행({exp_id_row}) 값: '{id_value}'")
                
                # 🆕 정규화: 모든 특수문자 제거
                id_value_clean = id_value.strip()
                # 🔥 추가: 콜론, 세미콜론, 점 등 모든 특수문자 제거
                import re
                id_value_clean = re.sub(r'[^A-Za-z0-9]', '', id_value_clean)
                
                # 🆕 숫자 → 알파벳 변환 (1 → I)
                if id_value_clean == '1':
                    id_value_clean = 'I'
                    print(f"    🔧 숫자 ID 보정: '1' → 'I'")
                elif id_value_clean == '0':
                    # 이전 컬럼 확인하여 O 또는 D 결정
                    pass
                
                # 단일 알파벳인지 확인
                if len(id_value_clean) == 1 and id_value_clean.isalpha():
                    has_experiment_id = True
                    experiment_id_value = id_value_clean.upper()
                    print(f"    ✅ 실험 ID '{experiment_id_value}' 발견! (원본: '{id_value}')")
                else:
                    print(f"    ❌ 단일 알파벳 아님 (정규화 후: '{id_value_clean}')")
            
            # ========== 🔥 2단계: 데이터 존재 여부 확인 ==========
            has_data = False
            data_count = 0
            found_rows = []
            
            for check_row_idx in range(exp_id_row, min(header_row + 20, len(table_matrix))):
                if check_row_idx in table_matrix:
                    row = table_matrix[check_row_idx]
                    
                    if col_idx in row:
                        cell_value = str(row[col_idx]).strip()
                        if cell_value and cell_value not in ['nan', 'None', '']:
                            data_count += 1
                            found_rows.append(check_row_idx)
                            if not has_data:
                                has_data = True
                        
                        # 처음 3개만 출력
                        if check_row_idx < exp_id_row + 3:
                            print(f"    행 {check_row_idx}: '{cell_value[:20] if len(cell_value) > 20 else cell_value}' → {bool(cell_value)}")
                    else:
                        if check_row_idx < exp_id_row + 3:
                            print(f"    행 {check_row_idx}: (키 없음)")
            
            print(f"    → has_data={has_data}, data_count={data_count}, found_rows={found_rows[:3]}...")
            
            # ========== 🔥 3단계: 조건 판단 ==========
            # 기존 조건 완화: name_col 바로 다음 컬럼도 실험 컬럼 가능성 고려
            if has_experiment_id and has_data and data_count > 0:
                experiment_cols.append(col_idx)
                print(f"    ✅ 실험 컬럼으로 추가! (ID: {experiment_id_value})")
            # 🆕 수정: name_col + 1 컬럼도 포함 (>= 대신 >)
            elif not has_experiment_id and data_count >= 5 and col_idx >= name_col + 1:  # 🔧 수정
                # 🆕 추가 검증: 알파벳 순서 확인
                # 이전/다음 컬럼과 순서가 맞으면 실험 컬럼으로 추가
                should_add = False
                
                # 이미 추가된 실험 컬럼이 있는 경우
                if experiment_cols:
                    last_exp_col = experiment_cols[-1]
                    # 연속된 컬럼이면 실험 컬럼일 가능성 높음
                    if col_idx == last_exp_col + 1:
                        should_add = True
                        print(f"    💡 이전 실험 컬럼과 연속: Col_{last_exp_col} → Col_{col_idx}")
                
                if should_add:
                    experiment_cols.append(col_idx)
                    print(f"    ✅ 실험 컬럼으로 추가! (ID 없지만 데이터 충분: {data_count}개)")
            else:
                # ✅ 추가: 제외 상세 이유
                print(f"    ❌ 제외됨")
                print(f"      - has_experiment_id: {has_experiment_id}")
                print(f"      - data_count: {data_count}")
                print(f"      - col_idx > name_col + 1: {col_idx} > {name_col + 1} = {col_idx > name_col + 1}")
                if not has_experiment_id and data_count < 5:
                    print(f"      → 사유: 실험 ID 없고 데이터 부족 ({data_count} < 5)")
                elif not has_experiment_id and col_idx <= name_col + 1:
                    print(f"      → 사유: 원료명 영역으로 추정")
        
        experiment_cols.sort()
        print(f"\n🧪 실험 컬럼 인덱스: {experiment_cols}")
        
        
        # 🆕 연속성 확인: 첫 컬럼 이전 + 중간 gap
        if len(experiment_cols) >= 1:
            print(f"\n🔍 실험 컬럼 연속성 확인 중...")
            missing_cols = []
            
            first_exp_col = experiment_cols[0]
            
            # 🆕 1단계: 첫 번째 실험 컬럼 이전 확인 (name_col 다음부터)
            if first_exp_col > name_col + 1:
                print(f"  💡 첫 실험 컬럼(Col_{first_exp_col}) 이전 확인")
                
                for check_col in range(name_col + 1, first_exp_col):
                    # 실험 ID 행에 값이 있는지 확인
                    if exp_id_row in table_matrix and check_col in table_matrix[exp_id_row]:
                        id_value = self._clean_checkbox_and_newline(str(table_matrix[exp_id_row][check_col]))
                        # 특수문자 제거
                        import re
                        id_value_clean = re.sub(r'[^A-Za-z0-9]', '', id_value.strip())
                        
                        # 빈 문자열이 아니면 후보
                        if id_value_clean or check_col == first_exp_col - 1:
                            print(f"    ⚠️ Col_{check_col} 누락 가능성 (ID: '{id_value}' → '{id_value_clean}')")
                            missing_cols.append(check_col)
            
            # 🆕 2단계: 기존 실험 컬럼 사이 gap 확인
            for i in range(len(experiment_cols) - 1):
                curr_col = experiment_cols[i]
                next_col = experiment_cols[i + 1]
                
                if next_col - curr_col > 1:
                    for missing_col in range(curr_col + 1, next_col):
                        print(f"    ⚠️ Col_{curr_col}과 Col_{next_col} 사이에 Col_{missing_col} 누락")
                        missing_cols.append(missing_col)
            
            # 누락 컬럼 추가
            if missing_cols:
                print(f"  🔧 누락 컬럼 추가: {missing_cols}")
                experiment_cols.extend(missing_cols)
                experiment_cols.sort()
                print(f"  ✅ 확장된 실험 컬럼: {experiment_cols}")

        print(f"\n🧪 최종 실험 컬럼 인덱스: {experiment_cols}")

        return {
            'phase_col': phase_col,
            'code_col': code_col,
            'name_col': name_col,
            'experiment_cols': experiment_cols
        }
            
    def _infer_missing_experiment_ids(self, experiment_cols: List[int], experiment_ids: Dict) -> Dict:
        """
        누락된 실험 ID를 주변 알파벳으로 추론
        
        예: P(col_9) - ?(col_10) - R(col_11) → Q로 추론
        
        개선사항:
        - 숫자 ID 보정 (0→D/O, 1→I)
        - UnboundLocalError 수정
        - 디버깅 로그 추가
        """
        import string
        
        sorted_cols = sorted(experiment_cols)
        result = experiment_ids.copy()
        alphabet = list(string.ascii_uppercase)
        
        print(f"\n🔍 누락된 실험 ID 추론 중...")
        
        # ✅ 추론 전 상태 출력
        print(f"  추론 전 매핑:")
        for col in sorted_cols:
            exp_id = experiment_ids.get(col, None)
            print(f"    Col_{col}: {exp_id if exp_id else '(없음)'}")
        
        # ========== 1단계: 숫자 ID 보정 ==========
        for col in sorted_cols:
            exp_id = result.get(col)
            
            # 🆕 'H-' 같은 경우 정규화
            if exp_id and '-' in exp_id:
                cleaned = exp_id.replace('-', '').replace('_', '').strip()
                if len(cleaned) == 1 and cleaned.isalpha():
                    result[col] = cleaned
                    print(f"  🔧 특수문자 제거: Col_{col} '{exp_id}' → '{cleaned}'")
                    exp_id = cleaned
            
            # 기존 숫자 ID 보정
            if exp_id == '0':
                idx = sorted_cols.index(col)
                if idx > 0:
                    prev_col = sorted_cols[idx-1]
                    prev_id = result.get(prev_col)
                    if prev_id == 'C':
                        result[col] = 'D'
                        print(f"  🔧 숫자 ID 보정: Col_{col} '0' → 'D' (C 다음)")
                    elif prev_id == 'N':
                        result[col] = 'O'
                        print(f"  🔧 숫자 ID 보정: Col_{col} '0' → 'O' (N 다음)")
            
            elif exp_id == '1':
                result[col] = 'I'
                print(f"  🔧 숫자 ID 보정: Col_{col} '1' → 'I'")
        
        # ========== 2단계: 누락된 ID 추론 ==========
        for i, col in enumerate(sorted_cols):
            # 🔧 먼저 정의
            prev_id = None
            next_id = None
            
            # 이미 ID가 있으면 건너뛰기
            if col in result and result[col]:
                continue
            
            # 이전 컬럼 ID 찾기
            if i > 0:
                prev_col = sorted_cols[i-1]
                if prev_col in result and result[prev_col]:
                    prev_id = result[prev_col]
            
            # 다음 컬럼 ID 찾기
            if i < len(sorted_cols) - 1:
                next_col = sorted_cols[i+1]
                if next_col in result and result[next_col]:
                    next_id = result[next_col]
            
            # 디버깅 로그
            print(f"  Col_{col} 추론:")
            print(f"    이전: Col_{sorted_cols[i-1] if i > 0 else 'N/A'} = {prev_id}")
            print(f"    다음: Col_{sorted_cols[i+1] if i < len(sorted_cols)-1 else 'N/A'} = {next_id}")
            
            # 🆕 추론 로직 (순차 우선)
            inferred_id = None
            
            # 이전 알파벳이 있는 경우 → 다음 알파벳
            if prev_id and len(prev_id) == 1 and prev_id.isalpha():
                prev_idx = alphabet.index(prev_id)
                inferred_id = alphabet[(prev_idx + 1) % 26]
                print(f"    💡 이전 ID 기반 추론: {prev_id} → {inferred_id}")
                
                # 🆕 다음 ID와 검증
                if next_id and len(next_id) == 1 and next_id.isalpha():
                    next_idx = alphabet.index(next_id)
                    expected_idx = alphabet.index(inferred_id)
                    
                    # 순서가 맞는지 확인
                    if expected_idx < next_idx or expected_idx == next_idx - 1:
                        print(f"    ✅ 순서 검증 통과: {inferred_id} < {next_id}")
                    else:
                        print(f"    ⚠️ 순서 불일치: {inferred_id} >= {next_id}")
                        inferred_id = f'Col_{col}'
            
            # 다음 알파벳만 있는 경우 → 이전 알파벳
            elif next_id and len(next_id) == 1 and next_id.isalpha():
                next_idx = alphabet.index(next_id)
                inferred_id = alphabet[(next_idx - 1) % 26]
                print(f"    💡 다음 ID 기반 추론: {next_id} → {inferred_id}")
            
            # 둘 다 없으면 fallback
            else:
                inferred_id = f'Col_{col}'
                print(f"    ⚠️ 추론 불가 → fallback")
            
            result[col] = inferred_id
            print(f"    → 최종: '{inferred_id}'")

        return result
        
    def _get_experiment_ids(self, table_matrix: Dict, exp_id_row: int, experiment_cols: List[int]) -> List[str]:
        """실험 ID 추출 (개선: 체크박스 및 줄바꿈 제거)"""
        if exp_id_row is None or exp_id_row not in table_matrix:
            print("⚠️ 실험 ID 행이 없습니다. 기본값 사용")
            return [f'{i+1}' for i in range(len(experiment_cols))]
        
        exp_ids = []
        row_data = table_matrix[exp_id_row]
        
        print(f"\n🔍 실험 ID 추출 상세:")
        
        for col_idx in experiment_cols:
            if col_idx in row_data:
                raw_value = str(row_data[col_idx]).strip()
                
                # 체크박스 및 줄바꿈 제거 (개선)
                cleaned_value = self._clean_checkbox_and_newline(raw_value)
                
                # X 변형을 X로 변환
                x_variants = ['×', '✕', '✗']
                if cleaned_value in x_variants:
                    cleaned_value = 'X'
                
                exp_id = cleaned_value.upper()
                
                # 단일 알파벳이면 사용
                if len(exp_id) == 1 and exp_id.isalpha():
                    exp_ids.append(exp_id)
                    print(f"  Col_{col_idx}: '{row_data[col_idx]}' → '{exp_id}' ✅")
                else:
                    fallback = f'{len(exp_ids)+1}'
                    exp_ids.append(fallback)
                    print(f"  Col_{col_idx}: '{row_data[col_idx]}' → '{fallback}' (fallback)")
            else:
                fallback = f'{len(exp_ids)+1}'
                exp_ids.append(fallback)
                print(f"  Col_{col_idx}: (없음) → '{fallback}' (fallback)")
        
        print(f"\n🧪 최종 실험 ID: {exp_ids}")
        return exp_ids
    
    def _merge_raw_materials(self, name_value: str, extra_cols: List) -> str:
        """
        Raw Materials 병합 (개선)
        
        Name 컬럼 외에 다른 컬럼에 원료명이 연장되어 있는 경우 병합
        """
        parts = []
        
        if name_value and name_value.strip():
            parts.append(name_value.strip())
        
        # 추가 컬럼 병합 (CODE, RAW MATERIALS 같은 헤더 제외)
        for extra_val in extra_cols:
            if extra_val and extra_val.strip():
                val = extra_val.strip()
                # 헤더나 불필요한 값 제외
                if val not in ['CODE', 'RAW MATERIALS', 'RAW_MATERIALS', 'MATERIAL', '원료', '원료명']:
                    parts.append(val)
        
        return ' '.join(parts)
    
    def _parse_and_clean_table(self, table) -> Dict:
        """테이블 파싱 및 정리"""
        print("\n🔧 테이블 전처리 시작...")
        
        table_matrix = {}
        for cell in table.cells:
            row_idx = cell.row_index
            col_idx = cell.column_index
            if row_idx not in table_matrix:
                table_matrix[row_idx] = {}
            table_matrix[row_idx][col_idx] = cell.content.strip()
        
            # ✅ 추가: 테이블 매트릭스 샘플 출력
        print("\n📊 테이블 매트릭스 샘플 (처음 5행):")
        for row_idx in range(min(5, len(table_matrix))):
            if row_idx in table_matrix:
                row_preview = {}
                for col_idx in sorted(table_matrix[row_idx].keys())[:8]:  # 처음 8개 컬럼만
                    value = table_matrix[row_idx][col_idx]
                    display_value = value[:20] if len(value) > 20 else value
                    row_preview[f"Col_{col_idx}"] = display_value
                print(f"  행 {row_idx}: {row_preview}")
            
        main_header_row, exp_id_row = self._find_header_rows(table_matrix)
        table_matrix = self._align_raw_materials_header(table_matrix, main_header_row)
        
        # 🎯 추가: RAW MATERIALS 헤더 정렬 전처리
        table_matrix = self._align_raw_materials_header(table_matrix, main_header_row)
        
        column_info = self._identify_columns(table_matrix, main_header_row, exp_id_row)
        
        if not column_info:
            return {'ingredients': [], 'experiment_columns': []}
        
        phase_col = column_info['phase_col']
        code_col = column_info['code_col']
        name_col = column_info['name_col']
        experiment_cols = column_info['experiment_cols']
        
        # 🎯 실험 ID 추출 (개선)
        experiment_ids = {}
        if exp_id_row is not None and exp_id_row in table_matrix:
            exp_row_data = table_matrix[exp_id_row]
            for exp_col in experiment_cols:
                if exp_col in exp_row_data:
                    raw_id = self._clean_checkbox_and_newline(exp_row_data[exp_col])
                    
                    # 🎯 X 변형 처리 (×, ✕, ✗ → X)
                    x_variants = ['×', '✕', '✗', '*']
                    if raw_id in x_variants:
                        raw_id = 'X'
                        print(f"  🔧 Col_{exp_col}: X 변형('{exp_row_data[exp_col]}') → 'X'로 변환")
                    
                    if raw_id and len(raw_id) <= 5:
                        experiment_ids[exp_col] = raw_id
        
        print(f"\n🧪 실험 ID 매핑 (초기): {experiment_ids}")
        
        # 🎯 누락된 실험 ID 추론 (Q 누락 등 해결)
        experiment_ids = self._infer_missing_experiment_ids(experiment_cols, experiment_ids)
        
        print(f"🧪 실험 ID 매핑 (최종): {experiment_ids}")
        
        # 🔥🔥🔥 여기부터 추가 🔥🔥🔥
        sorted_experiment_cols = sorted(experiment_cols)
        sorted_experiment_ids = [experiment_ids.get(col, f'Col_{col}') for col in sorted_experiment_cols]
        print(f"🧪 정렬된 실험 ID: {sorted_experiment_ids}")
        # 🔥🔥🔥 여기까지 추가 🔥🔥🔥
        
        # 성분 데이터 추출
        ingredients = []
        data_start_row = exp_id_row + 1 if exp_id_row else main_header_row + 1
        
        for row_idx in range(data_start_row, len(table_matrix)):
            if row_idx not in table_matrix:
                continue
            
            row_data = table_matrix[row_idx]
            
            phase = ''
            if phase_col in row_data:
                phase = self._clean_checkbox_and_newline(row_data[phase_col])
            
            code = ''
            if code_col in row_data:
                code = row_data[code_col].strip()
            
            if not code:
                continue
            
            # 원료명 추출 (name_col + name_col+1 병합)
            name_parts = []
            if name_col in row_data:
                name_val = row_data[name_col].strip()
                if name_val:
                    name_parts.append(name_val)
            
            # name_col + 1도 원료명으로 병합 (실험 컬럼이 아닌 경우)
            if name_col + 1 in row_data and (name_col + 1) not in experiment_cols:
                ext_val = row_data[name_col + 1].strip()
                if ext_val and ext_val not in ['nan', 'None', '']:
                    name_parts.append(ext_val)
            
            raw_materials = ' '.join(name_parts)
            
            ingredient = {
                'Phase': phase,
                'Code': code,
                'Raw_Materials': raw_materials
            }
            
            # 🔥 수정: experiment_cols → sorted_experiment_cols
            for exp_col in sorted_experiment_cols:
                exp_id = experiment_ids.get(exp_col, f'Col_{exp_col}')
                exp_value = ''
                
                if exp_col in row_data:
                    raw_value = row_data[exp_col]
                    
                    # 1단계: 체크박스 제거
                    exp_value = self._clean_checkbox_and_newline(raw_value)
                    
                    # 🆕 2단계: 정규화 (쉼표/콜론 → 점)
                    exp_value = self._normalize_experiment_value(exp_value)
                
                ingredient[exp_id] = exp_value
            
            ingredients.append(ingredient)
        
        # 보정 룰 적용
        print(f"\n🔧 보정 룰 적용 중...")
        # 🔥 수정: list(experiment_ids.values()) → sorted_experiment_ids
        ingredients = self._apply_data_correction_rules(
            ingredients, 
            sorted_experiment_ids
        )
        
        # 🔥 수정: list(experiment_ids.values()) → sorted_experiment_ids
        return {
            'ingredients': ingredients,
            'experiment_columns': sorted_experiment_ids
        }
        
    def _is_valid_ingredient_code(self, code: str) -> bool:
        """유효한 원료 코드 확인"""
        if not code or len(code) < 3:
            return False
        if code.isdigit():
            return False
        
        patterns = [r'^[A-Z0-9]{3,10}$', r'^[A-Z]{2,4}\d{3,6}$', r'^[A-Z]{3,6}$']
        for pattern in patterns:
            if re.match(pattern, code):
                return True
        return False
    
    def save_to_excel(self, formula_data: Dict, output_path: str):
        """Excel로 저장"""
        if not formula_data.get('ingredients'):
            print("❌ 저장할 데이터가 없습니다.")
            return
        
        print("\n💾 Excel 파일 생성 중...")
        
        df = pd.DataFrame(formula_data['ingredients'])
        base_cols = ['Phase', 'Code', 'Raw_Materials']
        
        original_order = formula_data.get('experiment_columns', [])
        exp_cols = [col for col in original_order if col in df.columns]
        
        df = df[base_cols + exp_cols]
        
        print(f"📊 DataFrame 생성: {len(df)}행 x {len(df.columns)}열")
        print(f"   컬럼 순서: {list(df.columns)}")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            workbook = writer.book
            worksheet = workbook.create_sheet('제형데이터', 0)
            
            info_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            info_font = Font(bold=True, size=10)
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            memo_fill = PatternFill(start_color='FFF9E6', end_color='FFF9E6', fill_type='solid')
            memo_font = Font(italic=True, color='999999', size=9)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            doc_info = [
                ['처방번호', formula_data.get('formula_number', '')],
                ['제품명', formula_data.get('product_name', '')],
                ['처방특성', formula_data.get('characteristics', '')]
            ]
            
            for row_idx, (label, value) in enumerate(doc_info, start=1):
                cell_label = worksheet.cell(row=row_idx, column=1, value=label)
                cell_label.fill = info_fill
                cell_label.font = info_font
                cell_label.border = thin_border
                cell_label.alignment = Alignment(horizontal='left', vertical='center')
                
                cell_value = worksheet.cell(row=row_idx, column=2, value=value)
                cell_value.border = thin_border
                cell_value.alignment = Alignment(horizontal='left', vertical='center')
                
                worksheet.merge_cells(start_row=row_idx, start_column=2, 
                                     end_row=row_idx, end_column=len(df.columns))
            
            for row_idx in range(4, 6):
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx, value='')
            
            header_row = 6
            for col_idx, col_name in enumerate(df.columns, start=1):
                cell = worksheet.cell(row=header_row, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            memo_row = 7
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=memo_row, column=col_idx, value='')
                cell.fill = memo_fill
                cell.font = memo_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            data_start_row = 8
            for df_row_idx, row_data in df.iterrows():
                excel_row = data_start_row + df_row_idx
                for col_idx, value in enumerate(row_data, start=1):
                    cell = worksheet.cell(row=excel_row, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            for col_idx in range(1, len(df.columns) + 1):
                max_length = 10
                col_letter = get_column_letter(col_idx)
                
                for row_idx in range(1, data_start_row + len(df)):
                    cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                
                worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)
            
            worksheet.row_dimensions[1].height = 25
            worksheet.row_dimensions[2].height = 25
            worksheet.row_dimensions[3].height = 25
            worksheet.row_dimensions[header_row].height = 30
            worksheet.row_dimensions[memo_row].height = 25
            
            worksheet.freeze_panes = 'D8'
            
            if 'raw_table' in formula_data and formula_data['raw_table'] is not None:
                raw_df = formula_data['raw_table']
                raw_df.to_excel(writer, sheet_name='원본데이터', index=False)
                
                raw_worksheet = writer.sheets['원본데이터']
                
                raw_header_fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
                raw_header_font = Font(bold=True, color='FFFFFF')
                
                for cell in raw_worksheet[1]:
                    cell.fill = raw_header_fill
                    cell.font = raw_header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        print(f"✅ Excel 저장 완료: {output_path}")
        print(f"   📊 시트1: 제형데이터 ({len(df)}행)")
        print(f"   📋 시트2: 원본데이터")


def main():
    """메인 실행"""
    print("="*80)
    print("🧴 화장품 제형 표 OCR 시스템 (예외 사례 보완 완성)")
    print("="*80)
    print("\n📋 적용된 보정 룰:")
    print("  RULE 1: 첫번째 실험 컬럼 공란 → '0'")
    print("  RULE 2: 'X', 'x', '-' → '0', 체크박스 제거")
    print("  RULE 3: 두번째 이후 컬럼 공란 → 이전 값 복사 (빈 컬럼 건너뛰기)")
    print("  RULE 4: Phase 공란 → 이전 Phase 상속")
    print("  RULE 5: 원료 코드 없는 행 삭제")
    print("  RULE 6: Phase 보정 (1→I, 0→O)")
    print("  RULE 7: 텍스트 → '0' (TO100 제외)")
    print("  RULE 8: 빈 시험 컬럼 감지 및 건너뛰기")
    print("\n🔧 예외 사례 처리:")
    print("  ✓ 체크박스 및 줄바꿈 제거 (Phase, Code, 실험 ID, 모든 값)")
    print("  ✓ 특수 숫자 형식 (2:0 → 2.0, :23.00 → 23.00)")
    print("  ✓ Raw Materials 자동 병합")
    print("="*80)
    
    ocr = KolmarCosmeticOCR()
    image_path = "스킨케어1팀_OCR추가자료x표시변환_250729_page_001_deskewed.png"
    formula_data = ocr.extract_cosmetic_formula_table(image_path)
    
    if formula_data and formula_data.get('ingredients'):
        print("\n" + "="*80)
        print("📊 추출 결과")
        print("="*80)
        print(f"📋 문서번호: {formula_data.get('formula_number')}")
        print(f"📦 제품명: {formula_data.get('product_name')}")
        print(f"🧴 원료 수: {len(formula_data['ingredients'])}개")
        print(f"🧪 실험 컬럼: {formula_data.get('experiment_columns')}")
        print("="*80)
        
        output_excel = f"{formula_data.get('formula_number', 'result')}_제형표.xlsx"
        ocr.save_to_excel(formula_data, output_excel)
        
        print("\n✅ 완료!")
    else:
        print("\n❌ 실패")


if __name__ == "__main__":
    main()