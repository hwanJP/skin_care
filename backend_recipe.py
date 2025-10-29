"""
제형 레시피 OCR 백엔드 (최소 구현)
기존 backend.py의 PDFProcessor만 재사용
"""

import io
import logging
import os
import tempfile  # 🔧 추가!

# ✅ 기존 backend에서 PDFProcessor만 import
from backend import PDFProcessor

# 🆕 Azure OCR
from azure_ai import KolmarCosmeticOCR

logger = logging.getLogger(__name__)

# ============================================
# 🆕 컬럼명 생성 함수 (C-1)
# ============================================
def _generate_experiment_column_name(index: int) -> str:
    """
    U부터 시작하는 Excel 스타일 컬럼명 생성
    index=0 → U, index=5 → Z, index=6 → AA, index=7 → AB, ...
    """
    excel_index = 21 + index  # U는 21번째 알파벳 (A=1, U=21)
    
    result = ''
    while excel_index > 0:
        excel_index -= 1
        result = chr(ord('A') + (excel_index % 26)) + result
        excel_index //= 26
    
    return result

def process_recipe_page(pdf_bytes: bytes, page_index: int) -> dict:
    """
    제형 레시피 페이지 처리 (간소화)
    """
    result = {
        'success': False,
        'data': [],
        'metadata': {},
        'experiment_columns': [],
        'message': ''
    }
    
    temp_image_path = None
    
    try:
        # 1. DRM 처리
        drm_success, processed_bytes, drm_message = PDFProcessor.process_drm_if_needed(pdf_bytes)
        if not drm_success:
            result['message'] = drm_message
            return result
        
        logger.info(f"📄 DRM 처리: {drm_message}")
        
        # 2. 이미지 렌더링
        img_bytes = PDFProcessor.render_page_image(processed_bytes, page_index, zoom=2.0)
        if not img_bytes:
            result['message'] = "이미지 렌더링 실패"
            return result
        
        # 3. 임시 파일 저장
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as temp_file:
            temp_image_path = temp_file.name
            temp_file.write(img_bytes)
        
        logger.info(f"💾 임시 이미지 저장: {temp_image_path}")
        
        # 4. Azure OCR
        ocr = KolmarCosmeticOCR()
        formula_data = ocr.extract_cosmetic_formula_table(temp_image_path)
        
        if not formula_data or not formula_data.get('ingredients'):
            result['message'] = "데이터 추출 실패"
            return result
        
        # 5. 결과 포맷팅
        result['success'] = True
        result['data'] = formula_data['ingredients']
        result['metadata'] = {
            'formula_number': formula_data.get('formula_number', ''),
            'product_name': formula_data.get('product_name', ''),
            'characteristics': formula_data.get('characteristics', '')
        }
        
        # 🔧 experiment_columns 처리
        exp_cols = formula_data.get('experiment_columns', [])
        
        # Col_4, Col_5 형태면 자동으로 알파벳 생성
        if exp_cols and all(col.startswith('Col_') for col in exp_cols):
            logger.warning(f"⚠️ 기본 컬럼명 감지: {exp_cols}")
            
            # ✅ 새로운 컬럼명 생성 함수 사용
            new_exp_cols = []
            for i, old_col in enumerate(exp_cols):
                new_col = _generate_experiment_column_name(i)
                new_exp_cols.append(new_col)
                
                # 데이터에서도 컬럼명 변경
                for ingredient in result['data']:
                    if old_col in ingredient:
                        ingredient[new_col] = ingredient.pop(old_col)
            
            exp_cols = new_exp_cols
            logger.info(f"✅ 컬럼명 자동 변환: {exp_cols}")
        
        # ✅ 빈 리스트면 원료 데이터에서 추출 (내부 필드 필터링 H-1)
        elif not exp_cols and result['data']:
            first_ingredient = result['data'][0]
            base_cols = ['Phase', 'Code', 'Raw_Materials']
            internal_cols = ['_corrections', '_is_separator']  # 🆕 내부 필드
            
            exp_cols = [
                col for col in first_ingredient.keys() 
                if col not in base_cols and col not in internal_cols
            ]
            logger.info(f"🔧 experiment_columns 자동 생성: {exp_cols}")
        
        result['experiment_columns'] = exp_cols
        
        # ============================================
        # 🆕 여분 컬럼 추가 (OCR 시점에 1회만)
        # ============================================
        if exp_cols:
            last_col = exp_cols[-1]
            
            if len(last_col) == 1:  # 단일 문자 (U~Z)
                next_col = chr(ord(last_col) + 1)
            else:  # AA, AB 등
                next_col = last_col[:-1] + chr(ord(last_col[-1]) + 1)
            
            exp_cols_with_extra = exp_cols + [next_col]
            
            # 데이터에도 빈 값 추가
            for ingredient in result['data']:
                ingredient[next_col] = ''
            
            result['experiment_columns'] = exp_cols_with_extra
            logger.info(f"✅ 여분 컬럼 추가: {next_col}")
        
        logger.info(f"✅ OCR 성공: {len(result['data'])}개 원료, 실험 컬럼: {exp_cols}")
        result['message'] = f"{len(formula_data['ingredients'])}개 원료 추출 완료"
        
        return result
        
    except Exception as e:
        logger.error(f"❌ 처리 오류: {e}")
        import traceback
        traceback.print_exc()
        result['message'] = str(e)
        return result
    
    finally:
        # 임시 파일 정리
        if temp_image_path and os.path.exists(temp_image_path):
            try:
                os.remove(temp_image_path)
                logger.info(f"🗑️ 임시 파일 삭제: {temp_image_path}")
            except Exception as e:
                logger.warning(f"⚠️ 임시 파일 삭제 실패: {e}")



class RecipeExcelSaver:
    """제형 레시피 Excel 저장 (단순화)"""
    
    def __init__(self, output_path: str):
        self.output_path = output_path
        
        if not os.path.exists(self.output_path):
            from openpyxl import Workbook
            wb = Workbook()
            
            wb.save(self.output_path)
            wb.close()
    
    def add_recipe_data(self, data, metadata, experiment_cols):
        """제형 데이터 추가"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            import pandas as pd
            
            if not data:
                return False
            
            sorted_data = sorted(data, key=lambda x: x.get('Phase', ''))
            
            # ✅ DataFrame 생성 및 _corrections 제거 (C-2)
            df = pd.DataFrame(sorted_data)
            
            # _corrections 컬럼 제거 (Excel에 출력 안 함)
            if '_corrections' in df.columns:
                df = df.drop(columns=['_corrections'])
            
            base_cols = ['Phase', 'Code', 'Raw_Materials']
            exp_cols = [col for col in experiment_cols if col in df.columns]
            df = df[base_cols + exp_cols]
            
            workbook = load_workbook(self.output_path)
            
            # ============================================
            # 🆕 첫 번째 저장 시 기본 'Sheet' 삭제
            # ============================================
            if 'Sheet' in workbook.sheetnames and len(workbook.sheetnames) == 1:
                default_sheet = workbook['Sheet']
                workbook.remove(default_sheet)
                logger.info("초기 기본 시트 삭제")
            
            # ============================================
            # 🆕 시트명 생성 로직 (처방번호 기반)
            # ============================================
            formula_number = metadata.get('formula_number', 'Recipe').strip()
            saved_sheet_name = metadata.get('saved_sheet_name', None)
            
            if saved_sheet_name:
                # 재편집: 기존 시트명 사용 (덮어쓰기)
                sheet_name = saved_sheet_name
                logger.info(f"재편집: 기존 시트명 사용 ({sheet_name})")
            else:
                # 첫 저장: 중복 체크 후 시트명 생성
                sheet_name = formula_number if formula_number else 'Recipe'
                original_name = sheet_name
                counter = 2
                
                while sheet_name in workbook.sheetnames:
                    sheet_name = f"{original_name}_{counter}"
                    counter += 1
                
                logger.info(f"새 시트명 생성: {sheet_name}")
            
            # ============================================
            # 시트 생성 또는 덮어쓰기
            # ============================================
            if sheet_name in workbook.sheetnames:
                # 재편집: 기존 시트 위치 유지하며 덮어쓰기
                old_index = workbook.sheetnames.index(sheet_name)
                del workbook[sheet_name]
                worksheet = workbook.create_sheet(title=sheet_name, index=old_index)
                logger.info(f"시트 덮어쓰기: {sheet_name} (위치: {old_index})")
            else:
                # 신규: 새 시트 생성
                worksheet = workbook.create_sheet(title=sheet_name)
                logger.info(f"새 시트 생성: {sheet_name}")
        
            
            # 스타일
            info_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            info_font = Font(bold=True, size=10)
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            separator_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
            # ============================================
            # 🆕 노란색 배경 (자동 보정된 함량 값용)
            # ============================================
            yellow_fill = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')
            # 🆕 메모 행 스타일
            memo_fill = PatternFill(start_color='FFF9E6', end_color='FFF9E6', fill_type='solid')
            memo_font = Font(italic=True, color='999999', size=9)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # 상단 정보 (1-3행)
            doc_info = [
                ['처방번호', metadata.get('formula_number', '')],
                ['제품명', metadata.get('product_name', '')],
                ['처방특성', metadata.get('characteristics', '')]
            ]
            
            for row_idx, (label, value) in enumerate(doc_info, start=1):
                cell_label = worksheet.cell(row=row_idx, column=1, value=label)
                cell_label.fill = info_fill
                cell_label.font = info_font
                cell_label.border = thin_border
                
                cell_value = worksheet.cell(row=row_idx, column=2, value=value)
                cell_value.border = thin_border
                worksheet.merge_cells(start_row=row_idx, start_column=2, 
                                    end_row=row_idx, end_column=len(df.columns))
            
            # 헤더 (6행)
            header_row = 6
            for col_idx, col_name in enumerate(df.columns, start=1):
                cell = worksheet.cell(row=header_row, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            # ============================================
            # 🆕 메모 행 (7행)
            # ============================================
            memo_row = 7
            memo_data = metadata.get('memo', {})
            
            for col_idx, col_name in enumerate(df.columns, start=1):
                cell = worksheet.cell(row=memo_row, column=col_idx)
                cell.value = memo_data.get(col_name, '')
                cell.fill = memo_fill
                cell.font = memo_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            logger.info(f"메모 행 작성 완료 (7행)")
            
            # ============================================
            # 🆕 데이터 (8행부터) - 스타일 적용
            # ============================================
            data_start_row = 8
            excel_row = data_start_row
            previous_phase = None
            
            for df_row_idx, ingredient in enumerate(sorted_data):
                current_phase = ingredient.get('Phase', '')
                
                # Phase 변경 시 빈 행 추가
                if previous_phase and current_phase != previous_phase:
                    for col_idx in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=excel_row, column=col_idx, value='')
                        cell.border = thin_border
                        cell.fill = separator_fill
                    excel_row += 1
                
                previous_phase = current_phase
                
                # ✅ 보정 플래그 가져오기 (sorted_data에서, DataFrame 아님)
                corrections = ingredient.get('_corrections', {})
                
                # 데이터 행 작성
                for col_idx, col_name in enumerate(df.columns, start=1):
                    value = ingredient.get(col_name, '')
                    cell = worksheet.cell(row=excel_row, column=col_idx)
                    cell.border = thin_border
                    
                    # Phase
                    if col_name == 'Phase':
                        cell.value = value
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Code
                    elif col_name == 'Code':
                        cell.value = value
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Raw_Materials
                    elif col_name == 'Raw_Materials':
                        cell.value = value
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # 실험 컬럼
                    elif col_name in exp_cols:
                        try:
                            numeric_value = float(value)
                            cell.value = numeric_value
                            cell.number_format = '0.0000'
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                        except (ValueError, TypeError):
                            cell.value = value
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # ✅ 자동 보정된 셀만 노란색 배경
                        if col_name in corrections:
                            if corrections[col_name] in ['filled_zero', 'copied']:
                                cell.fill = yellow_fill
                
                excel_row += 1
            
            # ============================================
            # 열 너비 조정
            # ============================================
            for col_idx in range(1, len(df.columns) + 1):
                max_length = 10
                col_letter = get_column_letter(col_idx)
                for row_idx in range(1, excel_row):
                    cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)
            
            # ============================================
            # 틀 고정 (D8 - 메모 행 다음부터)
            # ============================================
            worksheet.freeze_panes = 'D8'
            
            workbook.save(self.output_path)
            workbook.close()
            
            logger.info(f"💾 Excel 저장: {sheet_name} ({len(df)}개 원료)")
            # ============================================
            # 🆕 시트명 반환 (재편집 추적용)
            # ============================================
            return {'success': True, 'sheet_name': sheet_name}
            
        except Exception as e:
            logger.error(f"❌ Excel 저장 실패: {e}")
            import traceback
            traceback.print_exc()
            return {'success': False, 'sheet_name': None}
    
    def get_excel_bytes(self):
        """Excel 바이트 반환"""
        try:
            if os.path.exists(self.output_path):
                with open(self.output_path, 'rb') as f:
                    return f.read()
        except:
            pass
        return None
    
    def get_statistics(self):
        """통계 반환"""
        try:
            from openpyxl import load_workbook
            if os.path.exists(self.output_path):
                wb = load_workbook(self.output_path, read_only=True)
                sheet_count = len(wb.sheetnames)
                wb.close()
                
                file_size = os.path.getsize(self.output_path)
                return {
                    'test_sheets': sheet_count,
                    'file_size': file_size,
                    'file_size_mb': round(file_size / (1024 * 1024), 2)
                }
        except:
            pass
        
        return {'test_sheets': 0, 'file_size': 0, 'file_size_mb': 0}